from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import List


@dataclass(frozen=True)
class GapRow:
    number: str
    governance_area: str
    current_state: str
    gap_finding: str
    priority: str


def _strip_md_bold(text: str) -> str:
    return re.sub(r"\*\*(.*?)\*\*", r"\1", text).strip()


def _split_md_row(line: str) -> List[str]:
    """
    Splits a markdown table row like:
      | a | b | c |
    into ["a", "b", "c"] (trimmed).
    """
    raw = line.strip()
    if not raw.startswith("|") or not raw.endswith("|"):
        return []
    parts = [c.strip() for c in raw.strip("|").split("|")]
    return parts


def parse_gap_table(md_text: str) -> List[GapRow]:
    lines = [ln.rstrip("\n") for ln in md_text.splitlines()]

    # Find the header row starting with "| #"
    header_idx = None
    for i, ln in enumerate(lines):
        if ln.lstrip().startswith("| #") and "|" in ln:
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("Could not find markdown table header row starting with '| #'.")

    # Expect delimiter line next
    if header_idx + 2 >= len(lines):
        raise ValueError("Markdown table appears truncated.")

    data_lines = []
    for ln in lines[header_idx + 2 :]:
        if not ln.strip().startswith("|"):
            break
        # Skip delimiter rows if any
        if re.match(r"^\s*\|\s*-+\s*\|", ln):
            continue
        data_lines.append(ln)

    rows: List[GapRow] = []
    for ln in data_lines:
        cols = _split_md_row(ln)
        if len(cols) < 5:
            continue
        number, area, current_state, gap, priority = cols[:5]
        rows.append(
            GapRow(
                number=_strip_md_bold(number),
                governance_area=_strip_md_bold(area),
                current_state=_strip_md_bold(current_state),
                gap_finding=_strip_md_bold(gap),
                priority=_strip_md_bold(priority),
            )
        )

    if not rows:
        raise ValueError("No data rows parsed from markdown table.")
    return rows


def build_workbook(rows: List[GapRow], out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Gap Analysis"

    title = "Comprehensive Governance Gap Analysis"
    ws["A1"] = title
    ws.merge_cells("A1:E1")
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws["A2"] = (
        "Systematic audit of VaultEdge’s as-is operational state against standard IT control "
        "baselines, identifying deep technical and administrative vulnerabilities."
    )
    ws.merge_cells("A2:E2")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    headers = ["#", "Governance Area", "Current State at VaultEdge", "Gap / Finding", "Priority"]
    header_row = 4
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")  # slate-ish
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for r_idx, r in enumerate(rows, start=header_row + 1):
        ws.cell(row=r_idx, column=1, value=r.number)
        ws.cell(row=r_idx, column=2, value=r.governance_area)
        ws.cell(row=r_idx, column=3, value=r.current_state)
        ws.cell(row=r_idx, column=4, value=r.gap_finding)
        ws.cell(row=r_idx, column=5, value=r.priority)

    # Wrap + top align for body
    for row in ws.iter_rows(min_row=header_row + 1, max_row=header_row + len(rows), min_col=1, max_col=5):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Column widths
    widths = [5, 28, 55, 55, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze panes just below header
    ws.freeze_panes = ws["A5"]

    # Autofilter
    ws.auto_filter.ref = f"A{header_row}:E{header_row + len(rows)}"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> None:
    here = Path(__file__).resolve().parent
    md_path = here / "governance-gap-analysis.md"
    out_path = here / "governance-gap-analysis.xlsx"

    md_text = md_path.read_text(encoding="utf-8")
    rows = parse_gap_table(md_text)
    build_workbook(rows, out_path)
    print(f"Wrote: {out_path}")


if __name__ == "__main__":
    main()

